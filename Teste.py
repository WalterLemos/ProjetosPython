import pyautogui 
import pandas as pd
import time
import openpyxl

# Abre o arquivo Excel
workbook = openpyxl.load_workbook('C:\ProjetorPython\Publicacoes2.xlsx')

# Seleciona a planilha desejada
sheet = workbook['Planilha1']

# Obtém o número total de linhas na planilha
total_rows = sheet.max_row
# Define o índice inicial da linha
current_row =0 

while current_row >= total_rows:  
     NR_Processo = sheet.cell(row=2, column=1).value
     DT_Publicacao = sheet.cell(row=2, column=2).value
     Ato_Processual = sheet.cell(row=2, column=3).value
     Andamento = sheet.cell(row=2, column=4).value
     print("Valor 1:", NR_Processo)
     print("Valor 2:", DT_Publicacao)
     print("Valor 3:", Ato_Processual)
     print("Valor 4:", Andamento)
 # Incrementa o índice da linha
     current_row += 1
     # Fecha o arquivo Excel
workbook.close()




